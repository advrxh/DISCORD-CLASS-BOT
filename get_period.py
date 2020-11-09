import openpyxl
from user_defined import wb_name
import arrow
from sys import exit

wb = openpyxl.load_workbook('tt.xlsx')
sheet1 = wb['Sheet1']

current_day = str(arrow.now().format('dddd'))
current_time = str(arrow.now().format('hh:mm'))
current_meridian = str(arrow.now().format('A'))


def return_prd(prd, wday):
    if prd == 1:
        for row in range(1, sheet1.max_row + 1):
            if str(sheet1.cell(row=row, column=1).value).lower() == wday:
                current_prd = str(sheet1.cell(row=row, column=2).value).lower()
                current_prd_no = 1
                return current_prd, current_prd_no
                break

    if prd == 2:
        for row in range(1, sheet1.max_row + 1 ):
            if str(sheet1.cell(row=row, column=1).value).lower() == wday:
                current_prd = str(sheet1.cell(row=row, column=3).value).lower()
                current_prd_no = 2
                return current_prd, current_prd_no
                break

    if prd == 3:
        for row in range(1, sheet1.max_row + 1):
            if str(sheet1.cell(row=row, column=1).value).lower() == wday:
                current_prd = str(sheet1.cell(row=row, column=4).value).lower()
                current_prd_no = 3
                return current_prd, current_prd_no
                break


def check_day():
    for row in range(1, sheet1.max_row + 1):
        if str(sheet1.cell(row=row, column=1).value).lower() == current_day.lower():
            working_day = str(sheet1.cell(row=row, column=1).value).lower()
            return working_day
        elif str(sheet1.cell(row=row, column=1).value).lower() == 'sunday':
            false_day = 'sunday'
            return false_day
        else:
            print('', end='')


def check_prd(wday):
    running = True
    while running:
        for row in range(1, sheet1.max_row + 1):
            if str(sheet1.cell(row=row, column=1).value).lower() == wday and current_meridian == 'AM':
                if int(current_time[:2]) <= 8 and int(current_time[:2]) >= 7:
                    PERIOD = return_prd(1, wday=wday)
                    return PERIOD
                    running = False

                elif int(current_time[:2]) <= 10 and int(current_time[:2]) >= 9:
                    PERIOD = return_prd(2, wday=wday)
                    return PERIOD
                    running = False

                elif int(current_time[:2]) <= 12 and int(current_time[:2]) >= 11:
                    PERIOD = return_prd(3, wday=wday)
                    return PERIOD
                    running = False

                else:
                    print('no class time')
            elif str(sheet1.cell(row=row, column=1).value).lower() == wday and current_meridian != 'AM':
                print('it aint class time!!')
                running = False
                exit()


def period():
    wday = check_day()
    prd = check_prd(wday=wday)
    return prd


